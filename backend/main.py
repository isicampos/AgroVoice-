from fastapi import FastAPI, Request, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, RedirectResponse, JSONResponse
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from ia import transcribir_archivo
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from models import (
    Productor,
    NuevoProductor,
    Registro,
    NuevoRegistro,
    Usuario,
    NuevoUsuario,
    Login,
    Predio,
    NuevoPredio,
    Cuartel,
    NuevoCuartel
)

from database import (
    crear_bd,

    # PRODUCTORES
    guardar_productor,
    obtener_productores,
    eliminar_productor,
    actualizar_productor,

    # PREDIOS
    guardar_predio,
    obtener_predios,

    # CUARTELES
    guardar_cuartel,
    obtener_cuarteles,

    # REGISTROS
    guardar_registro,
    obtener_registros_bd,
    eliminar_registro,
    actualizar_registro,

    # USUARIOS
    crear_usuario,
    obtener_usuario,

    # DASHBOARD
    total_productores,
    total_predios,
    total_cuarteles,
    total_registros,
    ultimos_registros
)


# =========================================================
# CONFIGURACIÓN
# =========================================================

app = FastAPI(
    title="API de AgroVoice",
    description="API oficial de AgroVoice",
    version="1.0.0"
)


# =========================================================
# CREAR BASE DE DATOS
# =========================================================

crear_bd()


# =========================================================
# FRONTEND
# =========================================================

app.mount(
    "/static",
    StaticFiles(directory="../frontend"),
    name="static"
)


def page(nombre_archivo: str):

    return FileResponse(
        f"../frontend/{nombre_archivo}"
    )


# =========================================================
# PÁGINA PRINCIPAL
# =========================================================

@app.get("/")
def inicio():

    return page("landing.html")


# =========================================================
# LOGIN WEB
# =========================================================

@app.get("/login_web")
def login_web():

    return page("login.html")


# =========================================================
# REGISTRO DE CUENTA WEB
# =========================================================

@app.get("/registro_web")
def registro_web():

    return page("registro.html")


# =========================================================
# DASHBOARD WEB
# =========================================================

@app.get("/dashboard")
def dashboard(request: Request):

    usuario = request.cookies.get("usuario")

    if not usuario:

        return RedirectResponse(
            "/login_web"
        )

    return page("dashboard.html")


# =========================================================
# PRODUCTORES WEB
# =========================================================

@app.get("/productores_web")
def productores_web():

    return page("productores.html")


# =========================================================
# PREDIOS WEB
# =========================================================

@app.get("/predios_web")
def predios_web():

    return page("predios.html")


# =========================================================
# CUARTELES WEB
# =========================================================

@app.get("/cuarteles_web")
def cuarteles_web():

    return page("cuarteles.html")


# =========================================================
# REGISTROS WEB
# =========================================================

@app.get("/registros_web")
def registros_web():

    return page("registros.html")


# =========================================================
# REPORTES WEB
# =========================================================

@app.get("/reportes_web")
def reportes_web():

    return page("reportes.html")


# =========================================================
# CONFIGURACIÓN WEB
# =========================================================

@app.get("/configuracion_web")
def configuracion_web():

    return page("configuracion.html")
@app.get("/contacto_web")
def contacto_web():
    return FileResponse("../frontend/contacto.html")


# =========================================================
# API PRODUCTORES
# =========================================================

@app.get(
    "/productores",
    response_model=list[Productor]
)
def listar_productores():

    datos = obtener_productores()

    lista = []

    for fila in datos:

        lista.append(

            Productor(

                id=fila[0],

                nombre=fila[1],

                predio=fila[2],

                cultivo=fila[3]

            )

        )

    return lista


# =========================================================
# AGREGAR PRODUCTOR
# =========================================================

@app.post(
    "/productores",
    response_model=Productor
)
def agregar_productor(
    productor: NuevoProductor
):

    guardar_productor(

        productor.nombre,

        productor.predio,

        productor.cultivo

    )

    datos = obtener_productores()

    ultimo = datos[-1]

    return Productor(

        id=ultimo[0],

        nombre=ultimo[1],

        predio=ultimo[2],

        cultivo=ultimo[3]

    )


# =========================================================
# EDITAR PRODUCTOR
# =========================================================

@app.put("/productores/{id}")
def editar_productor(

    id: int,

    productor: NuevoProductor

):

    actualizar_productor(

        id,

        productor.nombre,

        productor.predio,

        productor.cultivo

    )

    return {

        "mensaje": "Productor actualizado"

    }


# =========================================================
# ELIMINAR PRODUCTOR
# =========================================================

@app.delete("/productores/{id}")
def borrar_productor(id: int):

    eliminar_productor(id)

    return {

        "mensaje": "Productor eliminado"

    }


# =========================================================
# API PREDIOS
# =========================================================

@app.get(
    "/predios",
    response_model=list[Predio]
)
def listar_predios():

    datos = obtener_predios()

    lista = []

    for fila in datos:

        lista.append(

            Predio(

                id=fila[0],

                nombre=fila[1],

                productor=fila[2],

                superficie=fila[3],

                region=fila[4],

                comuna=fila[5]

            )

        )

    return lista


# =========================================================
# AGREGAR PREDIO
# =========================================================

@app.post("/predios")
def agregar_predio(
    predio: NuevoPredio
):

    guardar_predio(

        predio.nombre,

        predio.productor_id,

        predio.superficie,

        predio.region,

        predio.comuna

    )

    return {

        "mensaje": "Predio creado"

    }


# =========================================================
# API CUARTELES
# =========================================================

@app.get(
    "/cuarteles",
    response_model=list[Cuartel]
)
def listar_cuarteles():

    datos = obtener_cuarteles()

    lista = []

    for fila in datos:

        lista.append(

            Cuartel(

                id=fila[0],

                nombre=fila[1],

                predio=fila[2],

                cultivo=fila[3],

                variedad=fila[4],

                superficie=fila[5]

            )

        )

    return lista


# =========================================================
# AGREGAR CUARTEL
# =========================================================

@app.post("/cuarteles")
def agregar_cuartel(
    cuartel: NuevoCuartel
):

    guardar_cuartel(

        cuartel.nombre,

        cuartel.predio_id,

        cuartel.cultivo,

        cuartel.variedad,

        cuartel.superficie

    )

    return {

        "mensaje": "Cuartel creado"

    }


# =========================================================
# API REGISTROS
# =========================================================

@app.get(
    "/registros",
    response_model=list[Registro]
)
def listar_registros():

    datos = obtener_registros_bd()

    lista = []

    for fila in datos:

        lista.append(

            Registro(

                id=fila[0],

                fecha=fila[1],

                productor=fila[2],

                predio=fila[3],

                cuartel=fila[4],

                cultivo=fila[5],

                labor=fila[6],

                transcripcion=fila[7]

            )

        )

    return lista


# =========================================================
# AGREGAR REGISTRO
# =========================================================

@app.post(
    "/registros",
    response_model=Registro
)
def agregar_registro(
    registro: NuevoRegistro
):

    guardar_registro(

        registro.fecha,

        registro.productor,

        registro.predio,

        registro.cuartel,

        registro.cultivo,

        registro.labor,

        registro.transcripcion

    )

    datos = obtener_registros_bd()

    ultimo = datos[0]

    return Registro(

        id=ultimo[0],

        fecha=ultimo[1],

        productor=ultimo[2],

        predio=ultimo[3],

        cuartel=ultimo[4],

        cultivo=ultimo[5],

        labor=ultimo[6],

        transcripcion=ultimo[7]

    )


# =========================================================
# EDITAR REGISTRO
# =========================================================

@app.put("/registros/{id_registro}")
def editar_registro(

    id_registro: int,

    registro: NuevoRegistro

):

    actualizar_registro(

        id_registro,

        registro.fecha,

        registro.productor,

        registro.predio,

        registro.cuartel,

        registro.cultivo,

        registro.labor,

        registro.transcripcion

    )

    return {

        "mensaje": "Registro actualizado"

    }


# =========================================================
# ELIMINAR REGISTRO
# =========================================================

@app.delete("/registros/{id_registro}")
def borrar_registro(
    id_registro: int
):

    eliminar_registro(

        id_registro

    )

    return {

        "mensaje": "Registro eliminado"

    }


# =========================================================
# DASHBOARD API
# =========================================================

@app.get("/dashboard_api")
def dashboard_api():

    # ---------------------------------
    # PRODUCTORES
    # ---------------------------------

    try:
        productores = total_productores()
    except Exception:
        productores = 0


    # ---------------------------------
    # PREDIOS
    # ---------------------------------

    try:
        predios = total_predios()
    except Exception:
        predios = 0


    # ---------------------------------
    # CUARTELES
    # ---------------------------------

    try:
        cuarteles = total_cuarteles()
    except Exception:
        cuarteles = 0


    # ---------------------------------
    # REGISTROS
    # ---------------------------------

    try:
        registros = total_registros()
    except Exception:
        registros = 0


    # ---------------------------------
    # ÚLTIMOS REGISTROS
    # ---------------------------------

    try:
        ultimos = ultimos_registros()
    except Exception:
        ultimos = []


    # ---------------------------------
    # RESPUESTA
    # ---------------------------------

    return {
        "productores": productores,
        "predios": predios,
        "cuarteles": cuarteles,
        "registros": registros,
        "ultimos": ultimos
    }


# =========================================================
# REGISTRAR USUARIO
# =========================================================

@app.post("/usuarios")
def registrar_usuario(
    usuario: NuevoUsuario
):

    crear_usuario(

        usuario.nombre,

        usuario.correo,

        usuario.password,

        usuario.rol

    )

    return {

        "mensaje":
            "Usuario creado correctamente"

    }


# =========================================================
# LOGIN API
# =========================================================

@app.post("/login")
def login(
    datos: Login
):

    usuario = obtener_usuario(
        datos.correo
    )

    if usuario is None:

        return {

            "ok": False,

            "mensaje":
                "Usuario no existe"

        }


    if usuario[3] != datos.password:

        return {

            "ok": False,

            "mensaje":
                "Contraseña incorrecta"

        }


    response = JSONResponse({

        "ok": True,

        "nombre": usuario[1],

        "rol": usuario[4]

    })


    response.set_cookie(

        key="usuario",

        value=usuario[2],

        httponly=True

    )


    return response


# =========================================================
# USUARIO ACTUAL
# =========================================================

@app.get("/usuario_actual")
def usuario_actual(
    request: Request
):

    usuario = request.cookies.get(
        "usuario"
    )

    if not usuario:

        return {

            "usuario": ""

        }


    return {

        "usuario": usuario

    }


# =========================================================
# TOTAL USUARIOS
# =========================================================

@app.get("/usuarios_total")
def usuarios_total():

    return {

        "total": 1

    }


# =========================================================
# CERRAR SESIÓN
# =========================================================

@app.get("/logout")
def logout():

    response = RedirectResponse(
        "/"
    )

    response.delete_cookie(
        "usuario"
    )

    return response
@app.get("/exportar_excel")
def exportar_excel():

    registros = obtener_registros_bd()

    wb = Workbook()
    ws = wb.active
    ws.title = "Registros Agrícolas"

    encabezados = [
        "Fecha",
        "Productor",
        "Predio",
        "Cuartel",
        "Cultivo",
        "Labor",
        "Descripción"
    ]

    ws.append(encabezados)

    # Estilo de encabezados
    relleno = PatternFill(
        fill_type="solid",
        fgColor="1B6D2B"
    )

    fuente = Font(
        color="FFFFFF",
        bold=True
    )

    borde = Border(
        left=Side(style="thin", color="D9E2DA"),
        right=Side(style="thin", color="D9E2DA"),
        top=Side(style="thin", color="D9E2DA"),
        bottom=Side(style="thin", color="D9E2DA")
    )

    for celda in ws[1]:

        celda.fill = relleno
        celda.font = fuente
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        celda.border = borde


    # Agregar registros
    for registro in registros:

        ws.append([
            registro[1],
            registro[2],
            registro[3],
            registro[4],
            registro[5],
            registro[6],
            registro[7]
        ])


    # Formato de las celdas
    for fila in ws.iter_rows():

        for celda in fila:

            celda.border = borde

            celda.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )


    # Ancho de columnas
    anchos = {
        "A": 15,
        "B": 25,
        "C": 25,
        "D": 18,
        "E": 18,
        "F": 25,
        "G": 55
    }

    for columna, ancho in anchos.items():

        ws.column_dimensions[columna].width = ancho


    # Filtro
    ws.auto_filter.ref = ws.dimensions

    # Congelar encabezado
    ws.freeze_panes = "A2"

    # Altura encabezado
    ws.row_dimensions[1].height = 28


    # Guardar en memoria
    archivo = BytesIO()

    wb.save(archivo)

    archivo.seek(0)


    return StreamingResponse(
        archivo,
        media_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            'attachment; filename="AgroVoice_Registros.xlsx"'
        }
    )
@app.get("/exportar_pdf")
def exportar_pdf():

    registros = obtener_registros_bd()

    archivo = BytesIO()

    documento = SimpleDocTemplate(
        archivo,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    estilos = getSampleStyleSheet()

    titulo = ParagraphStyle(
        "TituloAgroVoice",
        parent=estilos["Title"],
        alignment=TA_CENTER,
        fontSize=20,
        textColor=colors.HexColor("#145c28"),
        spaceAfter=8
    )

    subtitulo = ParagraphStyle(
        "SubtituloAgroVoice",
        parent=estilos["Normal"],
        alignment=TA_CENTER,
        fontSize=11,
        textColor=colors.HexColor("#68756d"),
        spaceAfter=20
    )

    elementos = []

    elementos.append(
        Paragraph(
            "AGROVOICE",
            titulo
        )
    )

    elementos.append(
        Paragraph(
            "Sistema Inteligente de Registro Agrícola",
            subtitulo
        )
    )

    elementos.append(
        Paragraph(
            "Reporte de registros agrícolas",
            estilos["Heading2"]
        )
    )

    elementos.append(Spacer(1, 12))


    datos = [
        [
            "Fecha",
            "Productor",
            "Predio",
            "Cuartel",
            "Cultivo",
            "Labor",
            "Descripción"
        ]
    ]


    for registro in registros:

        datos.append([
            str(registro[1] or ""),
            str(registro[2] or ""),
            str(registro[3] or ""),
            str(registro[4] or ""),
            str(registro[5] or ""),
            str(registro[6] or ""),
            str(registro[7] or "")
        ])


    tabla = Table(
        datos,
        repeatRows=1,
        colWidths=[
            65,
            100,
            90,
            75,
            75,
            95,
            220
        ]
    )


    tabla.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#145c28")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, 0),
                "CENTER"
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#d9e2da")
            ),

            (
                "FONTNAME",
                (0, 1),
                (-1, -1),
                "Helvetica"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                7
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#f5faf5")
                ]
            ),

            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                5
            ),

            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                5
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                6
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                6
            )

        ])
    )


    elementos.append(tabla)

    elementos.append(Spacer(1, 15))


    elementos.append(
        Paragraph(
            f"Total de registros: {len(registros)}",
            estilos["Normal"]
        )
    )


    documento.build(elementos)

    archivo.seek(0)


    return StreamingResponse(
        archivo,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            'attachment; filename="AgroVoice_Reporte.pdf"'
        }
    )
from fastapi import UploadFile, File
from ia import transcribir_archivo

@app.post('/transcribir')
async def transcribir(audio: UploadFile = File(...)):
    contenido = await audio.read()
    texto = transcribir_archivo(contenido)
    return {'transcripcion': texto}
# =========================================================
# TRANSCRIPCIÓN DE AUDIO
# =========================================================

@app.post("/transcribir")
async def transcribir(audio: UploadFile = File(...)):

    contenido = await audio.read()

    texto = transcribir_archivo(contenido)

    return {
        "transcripcion": texto
    }
